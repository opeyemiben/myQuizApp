from selenium import webdriver
from time import sleep
import time 
import pandas as pd
import sqlite3
import xlsxwriter
import requests
import openpyxl
import random
from bs4 import BeautifulSoup
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import webbrowser
# Specify the login credentials and the site URL
your_username = "349229877"
your_password = "Hyphea25"
your_username1 = "444187793"
your_password1 = "Hyphea25"
url = "https://1xbet.ng/en"

# Start a new browser window
firefox_path = r'c:\Program Files (x86)\Mozilla Firefox 3.6 Beta 3\firefox.exe'
webbrowser.register('firefox', None, webbrowser.BackgroundBrowser(firefox_path))
webbrowser.get('firefox').open_new_tab(url)
webbrowser.maximize_window()
time.sleep(2)

signin_btn = webbrowser.find_element_by_id('curLoginForm')
signin_btn.click()
time.sleep(1)
email_btn = webbrowser.find_element_by_id('auth_id_email')
email_btn.click()
time.sleep
password_field = webbrowser.find_element_by_id('auth-fom-password')
with open('password.txt','r') as x:
    password = x.read()
password_field.send_keys(password)
password_field.send_keys(Keys.RETURN)
time.sleep(3)
login_btn = webbrowser.find_element_by_name('auth_login-button')
login_btn.click()
print('You are logged in!')
time.sleep(4)

search_field = webbrowser.find_element_by_id('')
search_field.send_keys('')
search_loop = webbrowser.find_element_by_class_name('')
search_loop.click()
print('The search was made successfully!')

# Click the login dropdown
driver.find_element_by_id("curLoginForm").click()

# Fill in the login form and submit it
driver.find_element_by_id("auth_id_email").send_keys(your_username)
driver.find_element_by_id("auth-form-password").send_keys(your_password)
driver.find_element_by_class_name("auth__login-button").click()

# Click the login dropdown
driver_1.find_element_by_id("curLoginForm").click()
# Fill in the login form and submit it
driver_1.find_element_by_id("auth_id_email").send_keys(your_username1)
driver_1.find_element_by_id("auth-form-password").send_keys(your_password1)
driver_1.find_element_by_class_name("auth__login-button").click()

# Wait for the login process to complete
time.sleep(5)

# Navigate to a page of your choice on the site
driver.get("https://1xbet.ng/en/live/table-tennis")

# Navigate to a page of your choice on the site
driver_1.get("https://1xbet.ng/en/live/table-tennis")

# Print the account balance
Account_Balance = driver.find_element_by_class_name("top-b-acc__amount").text
print(Account_Balance)

# Find and click on the Masters section
driver.find_element_by_id("c-2064427").click()

# Find and click on the Masters section
driver_1.find_element_by_id("c-2064427").click()

# prompt the user to enter a name
name = input("Enter a name to search: ")

# find all instances of the name on the page
elements = driver.find_elements_by_xpath("//*[contains(text(),'" + name + "')]")
elements1 = driver_1.find_elements_by_xpath("//*[contains(text(),'" + name + "')]")

# click on the first element where the name is found
if len(elements) > 0:
    elements[0].click()
else:
    print("Name not found on page")
# wait for the page to load
time.sleep(5)

# Wait for the "Set" button to become clickable
driver.find_element_by_id("tabindex=0").click()
driver.find_element_by_id("tabindex=-1").click()
driver_1.find_element_by_id("tabindex=0").click()
driver_1.find_element_by_id("tabindex=-1").click()

# Make a GET request to the current URL
response = requests.get(driver.current_url)
response1 = requests.get(driver_1.current_url)

# Parse the HTML content of the page
soup = BeautifulSoup(response.content, 'html.parser')
 
# Retrieve the current URL
current_url = response.url
current_url1 = response1.url

# Print the current URL
print(current_url)
print(current_url1)

# click on the first element where the name is found
if len(elements) > 0:
    elements[0].click()
else:
    print("Name not found on page")
# wait for the page to load
time.sleep(5)

# repeat a set of commands until a condition is met
for i in range(5) :

    # extract table data
    td= driver.page.find_all ("td")
    columns  = [value.text.replace ("\n", "") for value in td]

    # Find the index of the first occurrence of "over"
    start_index = columns.index(['over', *['']*3])

    # Find the index of the first occurrence of "under 19.5"
    end_index = columns.index(['under 20', *['']*3])

    # Slice the list to extract the desired range
    columns = columns[start_index:end_index+1]

    # select every 4th item of the over and under list
    columns [0: ] [ : : 4]
    columns [1: ] [ : : 4]
    columns [2: ] [ : : 4]
    columns [3: ] [ : : 4]
    over = columns [1: ] [ : : 4]
    under = columns [2: ] [ : : 4]
    odd = over and under
    columns_names = ["over.text.strip", "odd.text.strip", "under.text.strip", "odd.text.strip"]
    dictionary = {"over" : over}
    for idx, key in enumerate(columns_names):
        dictionary [key]= columns [idx: ] [ : : 4] 
    df = pd.DataFrame(data = dictionary)
    print (df())

    def generate_excel(workbook_name: str, worksheet_name: str, headers_list: columns[str], data: columns):
        workbook_name = "latest-excel.xlsx" 
    # Load the workbook and worksheet
    workbook = load_workbook(workbook_name = "latest-excel.xlsx")
    worksheet = workbook['Sheet1']
    Sheet1 = "sheet" 
    # Input data into specific columns
    data = [(over.text.strip(), odd.text.strip(), under.text.strip(), odd.text.strip())]
    df_Over_text = "Over_text"
    df_stake_ver = "stake_ver"
    df_profit_loss = "profit/loss"
    df_stake_nder = "stake_nder"
    df_Under_text = "Under_text"
    for row_num, row_data in enumerate(data, start=1):
        worksheet.cell(row=row_num, column=1, value=row_data[0])
        worksheet.cell(row=row_num, column=2, value=row_data[1])
        worksheet.cell(row=row_num, column=9, value=row_data[2])
        worksheet.cell(row=row_num, column=8, value=row_data[3])

    # Set the value of cell B16 to Acct Balance
    Sheet1['B16'] = Account_Balance

    # Save the changes to the Excel file
    workbook.save('latest-excel.xlsx')  
    # Write the data to the worksheet
    for row in dataframe_to_rows(data, index=False, header=False):
        worksheet.append(row)

        # Save the changes to the workbook
        workbook.save('latest-excel.xlsx')

    generate_excel("latest-excel.xlsx", "Sheet1", ["Over_text", "stake_ver", "profit/loss", "stake_nder", "Under_text"], [ df_Over_text["Over_text"], df_stake_ver["stake_ver"], df_profit_loss["profit/loss"], df_stake_nder["stake_nder"], df_Under_text["Under_text"]])

    df = pd.read_excel("latest-excel.xlsx", "Sheet1")
    print(df['Over_text'])
    print(df['stake_ver'])
    print(df['profit/loss'])
    print(df['stake_nder'])
    print(df['Under_text'])

    # Assuming you have a DataFrame called df with columns 'stake_ver', 'stake_nder', 'profit/loss', 'Over_text' and 'Under_text'

    # Read the data from a CSV file
    df = pd.read_csv('my_data.csv')

    # Concatenate the relevant columns into a new DataFrame
    concatenated = pd.concat([df['Over_text'], df['stake_ver'], df['profit/loss'], df['stake_nder'], df['Under_text']], axis=1)

    # Create empty lists to store the stake_ver, stake_nder, Over_text and Under_text data for profits
    Over_text_list = []
    stake_ver_list = []
    stake_nder_list = []
    Under_text_list = []

    # Loop through each item in the concatenated DataFrame
    for index, row in concatenated.iterrows():
        profit = row['profit/loss']
        Over_text = row['Over_text']
        stake_ver = row['stake_ver']
        stake_nder = row['stake_nder']
        Under_text = row['Under_text']

    # define the profit threshold
    profit_threshold = 10%(('stake_ver' + 'stake_nder')-'profit/loss')

    # assuming you have already defined and populated the following lists:
    Over_text_list, stake_ver_list, stake_nder_list, Under_text_list
    # determine the number of iterations based on the length of the lists
    num_iterations = max(len(Over_text_list), len(stake_ver_list), len(Under_text_list), len(stake_nder_list))

    # iterate over the lists and execute the code block for each item that meets the profit threshold
    for i in range(num_iterations):
        if i < len(Over_text_list) and int(Over_text_list[i]) > profit_threshold:
            print(f"Total over.text for profits above {profit_threshold}: {Over_text_list[i]}")
            print(f"stake_ver.text for profits above {profit_threshold}: {stake_ver_list[i]}")
            print(f"stake_nder.text for profits above {profit_threshold}: {stake_nder_list[i]}")
            print(f"Total under.text for profits above {profit_threshold}: {Under_text_list[i]}")

        # Retracking 1xbet account to place bet
        driver.get(current_url)
    #For under/over integer conditions stake,don't stake  the odds if less than 2.0 odds
    def stake_bet(condition, odds):
        if (condition == "under integer" or condition == "over integer") and odds < 2.0:
            return None 
 # Skip the bet if odds are less than 2.0 for condition "under integer" or "over integer".
        else:
            return "Bet placed successfully"
        result = stake_bet("under integer", 1.5)
        if result is None:
            print("Skipping the bet.")
        else:
            print("Bet placed successfully.")

def place_bet(condition):
    # Replace this function with your actual betting logic This is just a placeholder
    if condition:
        return True
    else:
        return False

def main():
    conditions = [
        ("over 14", "under 13.5"),
        ("over 15", "under 14.5"),
        ("over 16", "under 15.5"),
        ("over 17", "under 16.5"),
        ("over 18", "under 17.5"),
        ("over 19", "under 18.5"),
        ("over 20", "under 19.5")
    ]
    
    # Shuffle the conditions randomly
    random.shuffle(conditions)
    
    for over, under in conditions:
        if place_bet(over) and place_bet(under):
            print("Bet placed on conditions:", over, "and", under)
            conditions.remove((over, under))  
            # Remove the placed condition pair
        break  
        # Stop betting after the first successful bet
        print("Betting completed.")
    if __name__ == '__main__':
        main()
    # place bet on Over
    if i < len(Over_text_list):
        driver.find_element_by_class("bet_type>Total over.text[0]").click()
        driver.find_element_by_id("bet_input").send_keys(stake_ver_list[i])
        time.sleep(5)
        driver.find_element_by_id("place_bet").click()

    # place bet on Under
    if i < len(Under_text_list):
        driver_1.find_element_by_class("bet_type>Total under.text[0]").click()
        driver_1.find_element_by_id("bet_input").send_keys(stake_nder_list[i])
        time.sleep(5)
        driver_1.find_element_by_id("place_bet").click()
# commands to repeat go here
i += 1


